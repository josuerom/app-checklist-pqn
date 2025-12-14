import os
from datetime import datetime
from openpyxl import load_workbook
from flask import current_app


class ExcelService:
    """Servicio para manejar operaciones con Excel"""

    @staticmethod
    def generar_excel(tipo: str, respuestas: dict, session_data: dict) -> tuple:
        """
        Generar archivo Excel con el checklist completado

        Args:
            tipo: Tipo de checklist
            respuestas: Diccionario con las respuestas {pregunta_id: valor}
            session_data: Datos de sesión del usuario

        Returns:
            tuple: (ruta_archivo, nombre_archivo)
        """
        from app.models.checklist_data import get_checklist

        config = get_checklist(tipo)
        if not config:
            raise ValueError(f"Checklist tipo '{tipo}' no encontrado")

        # Cargar plantilla
        plantilla_path = os.path.join(
            current_app.config["TEMPLATES_DIR"], config["plantilla"]
        )

        if not os.path.exists(plantilla_path):
            raise FileNotFoundError(f"Plantilla no encontrada: {plantilla_path}")

        wb = load_workbook(plantilla_path)
        ws = wb.active

        # Llenar respuestas en el Excel
        final_line = ExcelService._llenar_respuestas(ws, respuestas)

        # Agregar información de validación
        ExcelService._agregar_validacion(ws, final_line, session_data)

        # Generar nombre de archivo
        nombre_archivo = ExcelService._generar_nombre_archivo(config, session_data)

        # Guardar archivo
        output_path = os.path.join(current_app.config["OUTPUT_DIR"], nombre_archivo)

        wb.save(output_path)
        current_app.logger.info(f"Excel generado: {output_path}")

        return output_path, nombre_archivo

    @staticmethod
    def _llenar_respuestas(ws, respuestas: dict) -> int:
        """Llenar las respuestas en la hoja de trabajo"""
        final_line = None

        for row in range(1, ws.max_row + 1):
            celda_id = ws.cell(row=row, column=1).value

            if celda_id and str(celda_id).isdigit():
                pregunta_id = int(celda_id)
                if pregunta_id in respuestas:
                    # Columna 3 para las respuestas
                    ws.cell(row=row, column=3).value = respuestas[pregunta_id]

            final_line = row

        return final_line

    @staticmethod
    def _agregar_validacion(ws, final_line: int, session_data: dict):
        """Agregar línea de validación al final"""
        fecha = datetime.now().strftime("%d/%m/%Y")
        tecnico = session_data.get("tecnico", "Desconocido")
        validador = current_app.config["DEFAULT_VALIDATOR"]

        texto_validacion = (
            f"Fecha: {fecha}       "
            f"Técnico: {tecnico}       "
            f"Revisado por: {validador}"
        )

        ws.cell(row=final_line, column=1).value = texto_validacion

    @staticmethod
    def _generar_nombre_archivo(config: dict, session_data: dict) -> str:
        """Generar nombre de archivo sanitizado"""
        activo = session_data.get("activo_fijo", "SinAF")
        propietario = session_data.get("propietario", "SinPropietario").replace(
            " ", "-"
        )
        cargo = session_data.get("cargo", "SinCargo").replace(" ", "-")

        nombre_archivo = (
            f"Activo {activo} Checklist {config['empresa']} "
            f"{config['tipo']} {propietario} {cargo}.xlsx"
        )

        # Sanitizar nombre (eliminar caracteres no permitidos)
        caracteres_invalidos = r'\/:*?"<>|'
        for char in caracteres_invalidos:
            nombre_archivo = nombre_archivo.replace(char, "_")

        return nombre_archivo
