import pytest
import os
from datetime import datetime
from app import create_app
from app.services.excel_service import ExcelService
from app.services.file_service import FileService


@pytest.fixture
def app():
    """Fixture para crear la aplicación de prueba"""
    app = create_app()
    app.config["TESTING"] = True
    app.config["OUTPUT_DIR"] = "test_output"
    app.config["TEMPLATES_DIR"] = "templates_excel"
    return app


@pytest.fixture
def app_context(app):
    """Fixture para el contexto de la aplicación"""
    with app.app_context():
        yield app


class TestExcelService:
    """Tests para ExcelService"""

    def test_generar_nombre_archivo(self):
        """Test para generación de nombre de archivo"""
        config = {"empresa": "Proquinal", "tipo": "PC"}

        session_data = {
            "activo_fijo": "35456",
            "propietario": "Pepe Pérez",
            "cargo": "Desarrollador",
        }

        nombre = ExcelService._generar_nombre_archivo(config, session_data)

        assert "Activo 35456" in nombre
        assert "Proquinal" in nombre
        assert "PC" in nombre
        assert "Pepe-Pérez" in nombre
        assert "Desarrollador" in nombre
        assert nombre.endswith(".xlsx")

    def test_sanitizar_nombre_archivo(self):
        """Test para sanitización de nombres de archivo"""
        config = {"empresa": "Test/Company", "tipo": "PC:Test"}

        session_data = {
            "activo_fijo": "123*45",
            "propietario": "User<>Name",
            "cargo": "Dev|Ops",
        }

        nombre = ExcelService._generar_nombre_archivo(config, session_data)

        # Verificar que no contenga caracteres inválidos
        caracteres_invalidos = r'\/:*?"<>|'
        for char in caracteres_invalidos:
            assert char not in nombre

    def test_agregar_validacion(self, app_context):
        """Test para agregar línea de validación"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        session_data = {"tecnico": "Test Tecnico"}

        ExcelService._agregar_validacion(ws, 10, session_data)

        # Verificar que se agregó la validación
        valor_celda = ws.cell(row=10, column=1).value

        assert "Fecha:" in valor_celda
        assert "Técnico: Test Tecnico" in valor_celda
        assert "Revisado por:" in valor_celda

    def test_llenar_respuestas(self):
        """Test para llenar respuestas en Excel"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Simular estructura de plantilla
        ws.cell(row=1, column=1).value = 1
        ws.cell(row=2, column=1).value = 2
        ws.cell(row=3, column=1).value = 3

        respuestas = {1: "OK", 2: "N/A", 3: "PD"}

        final_line = ExcelService._llenar_respuestas(ws, respuestas)

        # Verificar respuestas
        assert ws.cell(row=1, column=3).value == "OK"
        assert ws.cell(row=2, column=3).value == "N/A"
        assert ws.cell(row=3, column=3).value == "PD"
        assert final_line >= 3


class TestFileService:
    """Tests para FileService"""

    def test_validar_plantilla_existente(self, app_context):
        """Test para validar plantilla existente"""
        # Crear plantilla temporal
        os.makedirs("templates_excel", exist_ok=True)
        plantilla_path = os.path.join("templates_excel", "test_plantilla.xlsx")

        from openpyxl import Workbook

        wb = Workbook()
        wb.save(plantilla_path)

        try:
            # Mock checklist con plantilla de prueba
            from app.models import checklist_data

            original_checklists = checklist_data.CHECKLISTS.copy()

            checklist_data.CHECKLISTS["test"] = {"plantilla": "test_plantilla.xlsx"}

            resultado = FileService.validar_plantilla("test")
            assert resultado is True

        finally:
            # Limpiar
            if os.path.exists(plantilla_path):
                os.remove(plantilla_path)
            checklist_data.CHECKLISTS = original_checklists

    def test_validar_plantilla_inexistente(self, app_context):
        """Test para validar plantilla inexistente"""
        resultado = FileService.validar_plantilla("tipo_inexistente")
        assert resultado is False

    def test_limpiar_archivos_antiguos(self, app_context):
        """Test para limpieza de archivos antiguos"""
        import time

        # Crear directorio temporal
        test_dir = "test_output"
        os.makedirs(test_dir, exist_ok=True)

        # Crear archivo de prueba
        test_file = os.path.join(test_dir, "test_old_file.txt")
        with open(test_file, "w") as f:
            f.write("test")

        # Modificar tiempo de modificación (hacer el archivo "antiguo")
        old_time = time.time() - (8 * 86400)  # 8 días atrás
        os.utime(test_file, (old_time, old_time))

        # Cambiar config temporalmente
        app = app_context
        original_dir = app.config["OUTPUT_DIR"]
        app.config["OUTPUT_DIR"] = test_dir

        try:
            # Ejecutar limpieza
            FileService.limpiar_archivos_antiguos(dias=7)

            # Verificar que el archivo fue eliminado
            assert not os.path.exists(test_file)

        finally:
            # Restaurar config y limpiar
            app.config["OUTPUT_DIR"] = original_dir
            if os.path.exists(test_dir):
                import shutil

                shutil.rmtree(test_dir)


class TestIntegration:
    """Tests de integración"""

    def test_flujo_completo_generacion_excel(self, app_context):
        """Test del flujo completo de generación de Excel"""

        # Preparar datos
        tipo = "pc"
        respuestas = {i: "OK" for i in range(1, 11)}  # 10 respuestas de prueba

        session_data = {
            "activo_fijo": "99999",
            "propietario": "Test User",
            "cargo": "Test Cargo",
            "tecnico": "Test Tecnico",
        }

        # Crear directorio de salida
        os.makedirs("test_output", exist_ok=True)

        try:
            # Cambiar config temporalmente
            app = app_context
            original_dir = app.config["OUTPUT_DIR"]
            app.config["OUTPUT_DIR"] = "test_output"

            # Generar Excel (solo si existe la plantilla)
            plantilla_path = os.path.join("templates_excel", "plantilla_pc.xlsx")

            if os.path.exists(plantilla_path):
                archivo_local, nombre_archivo = ExcelService.generar_excel(
                    tipo, respuestas, session_data
                )

                # Verificar que se generó el archivo
                assert os.path.exists(archivo_local)
                assert nombre_archivo.endswith(".xlsx")
                assert "Activo 99999" in nombre_archivo

                # Limpiar archivo generado
                if os.path.exists(archivo_local):
                    os.remove(archivo_local)
            else:
                pytest.skip("Plantilla no encontrada")

        finally:
            # Restaurar config
            app.config["OUTPUT_DIR"] = original_dir

            # Limpiar directorio de prueba
            if os.path.exists("test_output"):
                import shutil

                shutil.rmtree("test_output")


class TestHelpers:
    """Tests para funciones auxiliares"""

    def test_sanitize_filename(self):
        """Test para sanitización de nombres de archivo"""
        from app.utils.helpers import sanitize_filename

        nombre_sucio = "archivo/con\\caracteres:invalidos*?.txt"
        nombre_limpio = sanitize_filename(nombre_sucio)

        caracteres_invalidos = r'\/:*?"<>|'
        for char in caracteres_invalidos:
            assert char not in nombre_limpio

    def test_format_date(self):
        """Test para formateo de fechas"""
        from app.utils.helpers import format_date

        fecha = datetime(2025, 12, 12, 15, 30, 45)
        resultado = format_date(fecha)

        assert resultado == "12/12/2025"

    def test_create_response_data(self):
        """Test para creación de respuestas estructuradas"""
        from app.utils.helpers import create_response_data

        # Sin data
        response = create_response_data(True, "Operación exitosa")
        assert response["success"] is True
        assert response["message"] == "Operación exitosa"
        assert "data" not in response

        # Con data
        response = create_response_data(True, "OK", {"id": 123})
        assert response["success"] is True
        assert response["data"]["id"] == 123


# ===================================
# FIXTURES ADICIONALES
# ===================================


@pytest.fixture
def client(app):
    """Cliente de prueba para requests"""
    return app.test_client()


@pytest.fixture
def runner(app):
    """Runner para comandos CLI"""
    return app.test_cli_runner()


# ===================================
# CONFIGURACIÓN DE PYTEST
# ===================================


def pytest_configure(config):
    """Configuración inicial de pytest"""
    config.addinivalue_line("markers", "slow: marca tests que son lentos")
    config.addinivalue_line("markers", "integration: marca tests de integración")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
